import tkinter as tk
from tkinter import filedialog
import pandas as pd
import numpy as np
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
import matplotlib.pyplot as plt
from matplotlib.widgets import RadioButtons

# WAT Data Automation Tool
# Author: Rose Anne Lafuente
# Licensed Electronics Engineer | Product Engineer II | Python Automation
# Description:
#   Automates .wat-to-Excel workflows for semiconductor Wafer Acceptance Test (WAT) data.
#   Generates structured sheets and capability plots for yield analysis and reporting.
#   Features include:
#     - Dynamic wafer sheet naming (Wafer 1~N, where N is fetched from the .wat file)
#     - Per Unit Data, Per Wafer, and Summary sheets with audit-ready formatting
#     - Cp, Cpk, Cpk Hi, and Cpk Lo statistics for process capability evaluation
#     - Interactive histogram viewer with ±3σ normal curve overlay
#     - Scrollable GUI logs with success/error messages for transparency
#   Built with Python, Tkinter, OpenPyXL, Matplotlib, and NumPy.

class WATDataAutomation:
    def __init__(self, root):
        self.root = root
        self.root.title("WAT Data Automation")
        self.root.geometry("500x500")
        self.path_var = tk.StringVar()

        # Theme
        self.bg_color = "#f5f5f5"
        self.fg_color = "#222222"
        self.btn_bg = "#e0e0e0"
        self.btn_active = "#BEE395"
        self.root.configure(bg=self.bg_color)

        # Title
        title_frame = tk.Frame(self.root, bg=self.bg_color)
        title_frame.pack(pady=(10,0))
        tk.Label(title_frame, text="WAT Data Automation",
                 font=("Meiryo", 12, "bold"), fg="darkblue", bg=self.bg_color).pack(side="left")
        tk.Label(title_frame, text=" v1.0.0",
                 font=("Meiryo", 12, "italic"), fg="darkblue", bg=self.bg_color).pack(side="left")
        tk.Label(self.root, text="Developed by Rose Anne Lafuente | 2026",
                 font=("Arial", 7, "italic"), fg="gray", bg=self.bg_color).pack(pady=(0,10))

        self.create_file_selection_frame()
        self.create_action_buttons()
        self.create_status_box()
        self.create_exit_button()

    # --- Helper Functions ---
    def write_text_cell(self, ws, row, col, value):
        """Always store as text (used for WAF, SITE, labels)."""
        cell = ws.cell(row=row, column=col, value=str(value) if value not in (None, '') else "")
        cell.number_format = '@'   # Force Excel to treat as Text
        return cell
    
    def write_number_cell(self, ws, row, col, value, round_digits=3):
        try:
            if value not in (None, ''):
                num = float(value)
                if num.is_integer():
                    # Whole number → int
                    cell = ws.cell(row=row, column=col, value=int(num))
                else:
                    # Decimal → rounded float
                    rounded_val = round(num, round_digits)
                    cell = ws.cell(row=row, column=col, value=rounded_val)
                    cell.number_format = '0.000'
            else:
                ws.cell(row=row, column=col, value=value)
        except Exception:
            ws.cell(row=row, column=col, value=value)

    def autofit_columns(self, ws, min_col=1, min_row=1):
        for col in ws.iter_cols(min_col=min_col, min_row=min_row, max_row=ws.max_row):
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 3
            
    def apply_borders(self, ws, min_row=1, min_col=1):
        """Apply thin borders up to the last non-empty row, including merged cells."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Find the last row with any non-empty cell
        last_row = 0
        for row in range(1, ws.max_row + 1):
            if any(ws.cell(row=row, column=col).value not in (None, "")
                   for col in range(1, ws.max_column + 1)):
                last_row = row

        # Apply borders to normal cells
        for row in ws.iter_rows(min_row=min_row, max_row=last_row,
                                min_col=min_col, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                
    def find_param_rows(self, ws):
        """Find row indices for WAF, ID, SPEC HI, and SPEC LO in a worksheet."""
        waf_rows = [cell.row for row in ws.iter_rows(min_col=1, max_col=1)
                    for cell in row if str(cell.value).strip().upper() == "WAF"]
        id_rows = [cell.row for row in ws.iter_rows(min_col=1, max_col=1)
                   for cell in row if str(cell.value).strip().upper() == "ID"]
        spechi_rows = [cell.row for row in ws.iter_rows(min_col=1, max_col=1)
                       for cell in row if str(cell.value).strip().upper() == "SPEC HI"]
        speclo_rows = [cell.row for row in ws.iter_rows(min_col=1, max_col=1)
                       for cell in row if str(cell.value).strip().upper() == "SPEC LO"]

        return waf_rows, id_rows, spechi_rows, speclo_rows

    def build_param_mapping(self, ws, waf_rows, id_rows, spechi_rows, speclo_rows):
        """Build parameter mapping dictionary from WAF/ID/SPEC rows."""
        param_mapping = {}
        min_len = min(len(waf_rows), len(id_rows), len(spechi_rows), len(speclo_rows))
        for i in range(min_len):
            waf_row, id_row, hi_row, lo_row = waf_rows[i], id_rows[i], spechi_rows[i], speclo_rows[i]

            # FIX: use ws.cell() instead of ws[waf_row]
            param_headers = [ws.cell(row=waf_row, column=col).value
                             for col in range(3, ws.max_column + 1)]

            for idx, param in enumerate(param_headers):
                if param is None:
                    continue
                col = 3 + idx
                unit = ws.cell(row=id_row, column=col).value
                spec_hi = ws.cell(row=hi_row, column=col).value
                spec_lo = ws.cell(row=lo_row, column=col).value

                param_mapping[str(param).strip()] = (spec_hi, spec_lo, unit)

                if isinstance(spec_hi, (int, float)) and isinstance(spec_lo, (int, float)):
                    param_mapping[f"{param}_POS"] = (abs(spec_lo), abs(spec_hi), unit)

        return param_mapping

    def extract_site_values(self, ws, start_row=6, col="B"):
        """Extract unique SITE values from a worksheet starting at a given row."""
        site_values = []
        row = start_row
        while True:
            cell_val = ws[f"{col}{row}"].value
            if cell_val is None or str(cell_val).strip().upper() in ("AVERAGE", "STD", "SPEC"):
                break
            clean_site = str(cell_val).replace('-', '').strip()
            if clean_site and clean_site not in site_values:
                site_values.append(clean_site)
            row += 1
        return site_values
    
    def log_success(self, msg):
        self.show_status(f"✅ {msg}", color="black")

    def log_error(self, msg):
        self.show_status(f"❌ {msg}", color="red")

    # --- File selection ---
    def create_file_selection_frame(self):
        f = tk.LabelFrame(self.root, text="WAT File Selection", padx=10, pady=10,
                          bd=2, relief="groove", font=("Segoe UI", 10, "bold"))
        f.pack(fill="x", padx=15, pady=10)

        tk.Label(f, text="Select WAT File:").pack(side="left", padx=(0, 10), pady=5)
        tk.Entry(f, textvariable=self.path_var, width=40, bg="white", fg="black").pack(side="left", fill="x", expand=True)
        tk.Button(f, text="Browse", width=15,
                  command=self.browse_file,
                  bg=self.btn_bg, fg=self.fg_color, activebackground=self.btn_active).pack(side="right", padx=10)

    def browse_file(self):
        file_path = filedialog.askopenfilename(title="Select WAT File", filetypes=[("WAT files", "*.wat")])
        if file_path:
            self.path_var.set(file_path)
            self.excel_filename = file_path
            self.show_status(f"📂 Selected file: {file_path}")

    # --- Action buttons ---
    def create_action_buttons(self):
        action_frame = tk.Frame(self.root, bg=self.bg_color, height=70)
        action_frame.pack(pady=10, anchor="center")
        action_frame.pack_propagate(False)

        for col in range(3):
            action_frame.grid_columnconfigure(col, weight=1)

        tk.Button(action_frame, text="▶️  Run Automation", width=18,
                  command=self.run_automation,
                  bg="#92D050", fg=self.fg_color, activebackground=self.btn_active,
                  pady=2).grid(row=0, column=0, padx=15)

        tk.Button(action_frame, text="📊  Generate Summary", width=18,
                  command=self.run_summary,
                  bg="#4F81BD", fg="white", activebackground=self.btn_active,
                  pady=2).grid(row=0, column=1, padx=15)

        tk.Button(action_frame, text="📈  Histogram Plot", width=18,
                  command=self.run_histogram,
                  bg="#FFC000", fg=self.fg_color, activebackground=self.btn_active,
                  pady=2).grid(row=0, column=2, padx=15)

    # --- Status box ---
    def create_status_box(self):
        status_frame = tk.LabelFrame(self.root, text="", padx=10, pady=10)
        status_frame.pack(fill="both", expand=True, padx=15, pady=10)
        container = tk.Frame(status_frame)
        container.pack(fill="both", expand=True)

        self.status_box = tk.Text(container, height=10, wrap="word", bg="white", fg="black", state="disabled")
        self.status_vsb = tk.Scrollbar(container, orient="vertical", command=self.status_box.yview)
        self.status_hsb = tk.Scrollbar(container, orient="horizontal", command=self.status_box.xview)

        self.status_box.configure(yscrollcommand=self.status_vsb.set, xscrollcommand=self.status_hsb.set)
        self.status_box.grid(row=0, column=0, sticky="nsew")
        self.status_vsb.grid(row=0, column=1, sticky="ns")
        self.status_hsb.grid(row=1, column=0, sticky="ew")
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

    def show_status(self, message, color="black", clear=False):
        self.status_box.config(state="normal")
        if clear:
            self.status_box.delete("1.0", "end")
        self.status_box.insert("end", message + "\n")
        self.status_box.tag_add("lastline", "end-2l", "end-1c")
        self.status_box.tag_config("lastline", foreground=color)
        self.status_box.config(state="disabled")

    def clear_all(self):
        self.path_var.set("")
        self.show_status("", clear=True)

    def create_exit_button(self):
        frame = tk.Frame(self.root, bg=self.bg_color)
        frame.pack(fill="x", side="bottom", padx=15, pady=5)
        tk.Button(frame, text="EXIT", width=12, bg="#d32f2f", fg="white",
                  command=self.root.destroy).pack(side="right", pady=10)
        tk.Button(frame, text="Clear All", width=12, command=self.clear_all,
                  bg="#ffcccc", fg=self.fg_color, activebackground=self.btn_active).pack(side="right", padx=10)

    # --- Run Automation ---
    def run_automation(self):
        try:
            wat_file = self.path_var.get()
            if not wat_file:
                self.log_error("No file selected!")
                return

            excel_filename = wat_file.replace(".wat", ".xlsx")

            type_no, process, pcm_spec, lot_id, date, clean_qty = self.convert_to_excel(wat_file, excel_filename)
            self.add_per_unit_data(excel_filename, clean_qty)
            self.add_per_wafer_data(excel_filename, clean_qty, [], type_no, process, pcm_spec, lot_id, date)

            self.log_success(f"Automation complete. Deliverables saved to {excel_filename}")
        except Exception as e:
            self.log_error(f"Error during automation: {e}")
            
    # --- Convert WAT file to Excel ---

    def convert_to_excel(self, wat_file, excel_filename):
        try:
            # --- Step 1: Read full content ---
            with open(wat_file, 'r') as file:
                content_text = file.read()
                file.seek(0)
                content_lines = file.readlines()

            # --- Step 2: Extract metadata using regex ---
            wat_attached = "W.A.T DATA ATTACHED"
            type_no = re.search(r'TYPE NO\s*:(\S+)', content_text).group(1)
            process = re.search(r'PROCESS\s*:(\S+)', content_text).group(1)
            pcm_spec = re.search(r'PCM SPEC\s*:(\S+)', content_text).group(1)
            qty = re.search(r'QTY\s*:(.+)', content_text).group(1).strip()  # Includes 'pcs'
            lot_id = re.search(r'LOT ID\s*:(\S+)', content_text).group(1)
            date = re.search(r'DATE\s*:(\S+)', content_text).group(1)

            # Metadata cell positions
            data_to_excel = {
                (1, 5): wat_attached,
                (2, 1): "TYPE NO :", (2, 2): type_no,
                (2, 5): "PROCESS :", (2, 6): process,
                (2, 8): "PCM SPEC :", (2, 9): pcm_spec,
                (2, 11): "QTY :", (2, 12): qty,
                (3, 1): "LOT ID :", (3, 2): lot_id,
                (3, 5): "DATE :", (3, 6): date,
            }

            # --- Step 3: Extract table content ---
            header = content_lines[3].strip().split()
            data = [line.strip().split() for line in content_lines[4:] if line.strip()]
            df = pd.DataFrame(data, columns=header)

            # --- Step 4: Write to Excel ---
            wb = Workbook()
            ws = wb.active

            # Set the sheet name using qty (without 'pcs')
            clean_qty = qty.replace("pcs", "").strip()
            ws.title = f"Wafer 1~{clean_qty}"

            # Remove gridlines
            ws.sheet_view.showGridLines = False

            # Write metadata in specific cells
            for (row, col), value in data_to_excel.items():
                ws.cell(row=row, column=col, value=value)

            # Write DataFrame starting at row 4
            start_row = 4
            for r_idx, row_values in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
                if str(row_values[0]).strip().upper() == "AVERAGE":
                    self.write_text_cell(ws, r_idx, 1, "AVERAGE")
                    self.write_text_cell(ws, r_idx, 2, "")
                    for c_idx, value in enumerate(row_values[1:], start=3):
                        self.write_number_cell(ws, r_idx, c_idx, value)
                elif (
                    len(row_values) >= 3 and
                    isinstance(row_values[0], str) and
                    row_values[0].upper().startswith(("STD", "SPEC")) and
                    isinstance(row_values[1], str)
                ):
                    combined_label = f"{row_values[0]} {row_values[1]}"
                    self.write_text_cell(ws, r_idx, 1, combined_label)
                    for c_idx, value in enumerate(row_values[2:], start=3):
                        self.write_number_cell(ws, r_idx, c_idx, value)
                else:
                    for c_idx, value in enumerate(row_values, start=1):
                        if c_idx <= 2:  # WAF and SITE columns → text
                            self.write_number_cell(ws, r_idx, c_idx, value)
                        else:           # Parameters → numbers
                            self.write_number_cell(ws, r_idx, c_idx, value)

            # --- Step 5: Auto-fit column widths (from Column C, Row 4) ---
            self.autofit_columns(ws, min_col=3, min_row=4)

            # --- Step 6: Save file ---
            wb.save(excel_filename)
            self.log_success(f"The content has been successfully converted to an Excel file named {excel_filename}.")

            return type_no, process, pcm_spec, lot_id, date, clean_qty

        except Exception as e:
            self.log_error(f"Error in Convert to Excel step: {e}")
            return None
        
    # --- Add Per Unit Data ---
    def add_per_unit_data(self, excel_filename, clean_qty):
        wb = load_workbook(excel_filename)
        main_ws = wb[wb.sheetnames[0]]

        # Metadata
        type_no = main_ws["B2"].value
        process = main_ws["F2"].value
        pcm_spec = main_ws["I2"].value
        lot_id = main_ws["B3"].value
        date = main_ws["F3"].value

        # Extract SITE values using helper
        site_values = self.extract_site_values(main_ws, start_row=6, col="B")

        # Wafer IDs
        wafer_ids = []
        for wafer_num in range(1, int(clean_qty) + 1):
            wafer_label = f"TT_#{str(wafer_num).zfill(2)}"
            for site in site_values:
                wafer_ids.append((wafer_label, site))

        # Create sheet
        unit_ws = wb.create_sheet(title="per Unit Data")
        headers = ["TYPE", "PROCESS", "SPEC", "LOT", "DATE", "Wafer", "Site"]
        unit_ws.append(headers)
        for wafer, site in wafer_ids:
            site_number = site_values.index(site) + 1
            unit_ws.append([type_no, process, pcm_spec, lot_id, date, wafer, site_number])

        # Parameter blocks
        waf_rows, id_rows, spechi_rows, speclo_rows = self.find_param_rows(main_ws)

        current_header_col = 8
        for waf_row in waf_rows:
            # FIX: iterate over the row tuple correctly
            param_headers = [cell.value for cell in main_ws[waf_row] if cell.column >= 3]

            # Write headers
            for i, header in enumerate(param_headers):
                unit_ws.cell(row=1, column=current_header_col + i, value=header)

            start_value_row = waf_row + 2
            current_data_row = 2
            while True:
                label = main_ws[f"A{start_value_row}"].value
                if label in (None, "", "AVERAGE", "STD", "SPEC"):
                    break
                for i, header in enumerate(param_headers):
                    value = main_ws.cell(row=start_value_row, column=3 + i).value
                    self.write_number_cell(unit_ws, current_data_row, current_header_col + i, value)
                start_value_row += 1
                current_data_row += 1
            current_header_col += len(param_headers)

        # Add _POS columns
        param_headers = []
        col = 8
        while True:
            header = unit_ws.cell(row=1, column=col).value
            if header is None or header == "":
                break
            param_headers.append((col, header))
            col += 1
        next_col = col
        created_pos_columns = {}
        for col_idx, header in param_headers:
            for row in range(2, unit_ws.max_row + 1):
                val = unit_ws.cell(row=row, column=col_idx).value
                if isinstance(val, (int, float)) and val < 0:
                    pos_header = f"{header}_POS"
                    if pos_header not in created_pos_columns:
                        unit_ws.cell(row=1, column=next_col, value=pos_header)
                        created_pos_columns[pos_header] = next_col
                        next_col += 1
                    unit_ws.cell(row=row, column=created_pos_columns[pos_header], value=abs(val))

        # Formatting
        unit_ws.sheet_view.showGridLines = False
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        header_font = Font(bold=True)
        for cell in unit_ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        self.apply_borders(unit_ws)
        self.autofit_columns(unit_ws)

        wb.save(excel_filename)
        self.show_status(f"New 'Per Unit Data' sheet with formatting created.")

    # --- Add Per Wafer Data ---
    def add_per_wafer_data(self, excel_filename, clean_qty, site_values, type_no, process, pcm_spec, lot_id, date):
        wb = load_workbook(excel_filename)
        main_ws = wb["per Unit Data"]

        # Collect parameters
        Param = [main_ws.cell(row=1, column=col).value
                 for col in range(8, main_ws.max_column + 1)
                 if main_ws.cell(row=1, column=col).value]

        # Extract SITE values using helper
        data_ws = wb[wb.sheetnames[0]]
        site_values = self.extract_site_values(data_ws, start_row=6, col="B")

        site_start_col = 8  # Column H
        site_end_col = site_start_col + len(site_values) - 1
        
        per_Wafer_ws = wb.create_sheet(title="per Wafer")
        # Header row 1: General headers + merged 'Site' + stat headers
        headers_row1 = ["TYPE", "PROCESS", "SPEC", "LOT", "DATE", "WAF_ID", "Parameter"]
        headers_row1 += ["Site"] + [""] * (len(site_values) - 1)
        headers_row1 += ["AVERAGE", "STD_DEV", "SPEC HI", "SPEC LO", "Unit"]
        per_Wafer_ws.append(headers_row1)

        for idx, site in enumerate(site_values, start=8):  # site columns start at col 8
            self.write_number_cell(per_Wafer_ws, row=2, col=idx, value=site)
        # Merge metadata headers vertically from A1 to G1
        for col_num in range(1, 8):
            col_letter = get_column_letter(col_num)
            per_Wafer_ws.merge_cells(f"{col_letter}1:{col_letter}2")
            per_Wafer_ws[f"{col_letter}1"].alignment = Alignment(horizontal="center", vertical="center")
            
        # Merge 'Site' group
        
        if site_values:  # only merge if we have sites
            site_end_col = site_start_col + len(site_values) - 1
            self.site_end_col = site_end_col
            site_merge_range = f"{get_column_letter(site_start_col)}1:{get_column_letter(site_end_col)}1"
            per_Wafer_ws.merge_cells(site_merge_range)
            per_Wafer_ws.cell(row=1, column=site_start_col).value = "Site"
            per_Wafer_ws.cell(row=1, column=site_start_col).alignment = Alignment(horizontal="center", vertical="center")

        # Merge stat headers vertically from row 1 to row 2
        stat_headers = ["AVERAGE", "STD_DEV", "SPEC HI", "SPEC LO", "Unit"]
        stat_start_col = site_end_col + 1
        for i, header in enumerate(stat_headers):
            col_index = stat_start_col + i
            col_letter = get_column_letter(col_index)
            per_Wafer_ws.merge_cells(f"{col_letter}1:{col_letter}2")
            per_Wafer_ws[f"{col_letter}1"].value = header
            per_Wafer_ws[f"{col_letter}1"].alignment = Alignment(horizontal="center", vertical="center")

        # Add wafer rows
        for wafer_num in range(1, int(clean_qty) + 1):
            wafer_id = f"TT_#{str(wafer_num).zfill(2)}"
            for param in Param:
                per_Wafer_ws.append([type_no, process, pcm_spec, lot_id, date, wafer_id, param])
    
        # Copy and transpose values from "per Unit Data"
        unit_ws = wb["per Unit Data"]
        wafer_col = 6  # Column F (Wafer)
        param_start_col = 8  # Column H
        param_end_col = unit_ws.max_column

        # Get all unique wafer IDs
        wafer_ids = []
        for row in range(2, unit_ws.max_row + 1):
            wafer_val = unit_ws.cell(row=row, column=wafer_col).value
            if wafer_val and wafer_val not in wafer_ids:
                wafer_ids.append(wafer_val)

        for wafer in wafer_ids:
            wafer_rows = [row for row in range(2, unit_ws.max_row + 1)
                          if unit_ws.cell(row=row, column=wafer_col).value == wafer]
            if not wafer_rows:
                continue
            values_matrix = [[unit_ws.cell(row=row, column=col).value
                              for col in range(param_start_col, param_end_col + 1)]
                             for row in wafer_rows]
            if values_matrix:
                transposed = list(map(list, zip(*values_matrix)))
                for row_idx in range(3, per_Wafer_ws.max_row + 1):
                    if per_Wafer_ws.cell(row=row_idx, column=6).value == wafer:
                        for t_row_idx, t_row in enumerate(transposed):
                            for t_col_idx, val in enumerate(t_row):
                                cell = per_Wafer_ws.cell(row=row_idx + t_row_idx,
                                                         column=8 + t_col_idx,
                                                         value=round(val, 3) if isinstance(val, (int, float)) else val)
                                if isinstance(val, (int, float)):
                                    cell.number_format = '0.000'
                        break
                    
        # Find last row with a parameter
        last_param_row = 0
        for row in range(3, per_Wafer_ws.max_row + 1):
            if per_Wafer_ws.cell(row=row, column=7).value not in (None, ""):
                last_param_row = row

        # Add formulas only up to that row
        for row in range(3, last_param_row + 1):
            avg_col = site_end_col + 1
            stdev_col = site_end_col + 2

            avg_cell = per_Wafer_ws.cell(row=row, column=avg_col)
            stdev_cell = per_Wafer_ws.cell(row=row, column=stdev_col)

            avg_cell.value = f"=AVERAGE({get_column_letter(site_start_col)}{row}:{get_column_letter(site_end_col)}{row})"
            stdev_cell.value = f"=STDEV({get_column_letter(site_start_col)}{row}:{get_column_letter(site_end_col)}{row})"

            avg_cell.number_format = '0.000'
            stdev_cell.number_format = '0.000'
            
        # --- Inject Spec Mapping ---
        waf_rows, id_rows, spechi_rows, speclo_rows = self.find_param_rows(data_ws)
        param_mapping = self.build_param_mapping(data_ws, waf_rows, id_rows, spechi_rows, speclo_rows)

        # Fill Spec HI, Spec LO, Unit columns
        for row in range(2, per_Wafer_ws.max_row + 1):
            param_name = per_Wafer_ws.cell(row=row, column=7).value
            if param_name in param_mapping:
                spec_hi, spec_lo, unit = param_mapping[param_name]
                per_Wafer_ws.cell(row=row, column=site_end_col + 3, value=spec_hi)
                per_Wafer_ws.cell(row=row, column=site_end_col + 4, value=spec_lo)
                per_Wafer_ws.cell(row=row, column=site_end_col + 5, value=unit)

        # --- Format Header Rows of per Wafer sheet ---
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        header_font = Font(bold=True)

        # Apply to row 1 and row 2
        for row_idx in [1, 2]:
            for cell in per_Wafer_ws[row_idx]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
        # Auto-fit and borders
        per_Wafer_ws.sheet_view.showGridLines = False
        self.autofit_columns(per_Wafer_ws)
        self.apply_borders(per_Wafer_ws)
        wb.save(excel_filename)
        self.log_success(f"Per Wafer sheet with Spec Mapping added.")

    # --- Generate Statistical Summary ---
    def run_summary(self):
        excel_filename = self.path_var.get().replace(".wat", ".xlsx")
        if not excel_filename:
            self.show_status("⚠️ No Excel file selected yet.", color="red")
            return

        try:
            wb = load_workbook(excel_filename)
        except Exception as e:
            self.show_status(f"❌ Could not open {excel_filename}. Error: {e}", color="red")
            return

        unit_ws = wb["per Unit Data"]
        per_Wafer_ws = wb["per Wafer"]
        data_ws = wb[wb.sheetnames[0]]
        lot_id = data_ws["B3"].value or "UNKNOWN"

        # --- Detect site_end_col dynamically from row 2 ---
        site_start_col = 8  # Column H
        site_values = []
        for col in range(site_start_col, per_Wafer_ws.max_column + 1):
            val = per_Wafer_ws.cell(row=2, column=col).value
            if val not in (None, ""):
                site_values.append(val)
        site_end_col = site_start_col + len(site_values) - 1

        # --- Define styles ---
        fill_green = PatternFill(start_color="FF00823B", end_color="FF00823B", fill_type="solid")
        fill_blue  = PatternFill(start_color="FF156082", end_color="FF156082", fill_type="solid")
        white_bold = Font(bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

        # --- Create summary sheet ---
        param_map_ws = wb.create_sheet(title=f"{lot_id}_TTTT_WAT_Summary")
        param_map_ws.sheet_view.showGridLines = False

        # Merge headers
        param_map_ws.merge_cells("B2:B3")
        param_map_ws.merge_cells("C2:C3")
        param_map_ws.merge_cells("D2:D3")
        param_map_ws.merge_cells("E2:I2")

        # Title cell
        param_map_ws["E2"] = f"{lot_id}_TTTT_Summary"
        param_map_ws["E2"].alignment = Alignment(horizontal="center", vertical="center")
        param_map_ws["E2"].fill = fill_green
        param_map_ws["E2"].font = white_bold

        # Header labels
        param_map_ws["B2"] = "Parameter"
        param_map_ws["C2"] = "SPEC HI"
        param_map_ws["D2"] = "SPEC LO"
        param_map_ws["E3"] = "MEAN"
        param_map_ws["F3"] = "STDEV"
        param_map_ws["G3"] = "CPK"
        param_map_ws["H3"] = "CPK Hi"
        param_map_ws["I3"] = "CPK Lo"

        # Style B–D headers
        for row in [2, 3]:
            for col in range(2, 5):
                cell = param_map_ws.cell(row=row, column=col)
                cell.fill = fill_blue
                cell.font = white_bold
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style E–I headers
        for col in range(5, 10):
            cell = param_map_ws.cell(row=3, column=col)
            cell.fill = fill_green
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # --- Build mapping of Parameter : SPEC HI : SPEC LO ---
        param_map = {}
        for row in range(3, per_Wafer_ws.max_row + 1):
            param = per_Wafer_ws.cell(row=row, column=7).value
            spec_hi = per_Wafer_ws.cell(row=row, column=site_end_col + 3).value
            spec_lo = per_Wafer_ws.cell(row=row, column=site_end_col + 4).value

            if param:
                param = str(param).strip()
                if param not in param_map:
                    param_map[param] = {"SPEC HI": spec_hi, "SPEC LO": spec_lo}

        # --- Compute MEAN and STDEV from per Unit Data ---
        last_row = unit_ws.max_row
        for col in range(8, unit_ws.max_column + 1):  # start at H
            param_name = unit_ws.cell(row=1, column=col).value
            if not param_name:
                continue

            values = []
            for row in range(2, last_row + 1):
                val = unit_ws.cell(row=row, column=col).value
                if val is not None:
                    try:
                        values.append(float(val))
                    except ValueError:
                        pass

            if values:
                arr = np.array(values, dtype=float)
                mean_val = round(float(np.mean(arr)), 3)
                stdev_val = round(float(np.std(arr, ddof=1)), 3)  # STDEV.S
                if param_name.strip() in param_map:
                    param_map[param_name.strip()]["MEAN"] = mean_val
                    param_map[param_name.strip()]["STDEV"] = stdev_val

                    # --- Compute CPK metrics ---
                    spec_hi = param_map[param_name.strip()]["SPEC HI"]
                    spec_lo = param_map[param_name.strip()]["SPEC LO"]

                    if spec_hi is not None and spec_lo is not None and stdev_val != 0:
                        cpk1 = round((spec_hi - mean_val) / (3 * stdev_val), 3)
                        cpk2 = round((mean_val - spec_lo) / (3 * stdev_val), 3)
                        param_map[param_name.strip()]["CPK"] = min(cpk1, cpk2)
                        param_map[param_name.strip()]["CPK Hi"] = max(cpk1, cpk2)
                        param_map[param_name.strip()]["CPK Lo"] = min(cpk1, cpk2)
                    else:
                        param_map[param_name.strip()]["CPK"] = "na"
                        param_map[param_name.strip()]["CPK Hi"] = "na"
                        param_map[param_name.strip()]["CPK Lo"] = "na"

        # --- Paste into summary sheet ---
        row_num = 4
        for param, specs in param_map.items():
            for col_idx, key in enumerate(
                ["Parameter", "SPEC HI", "SPEC LO", "MEAN", "STDEV", "CPK", "CPK Hi", "CPK Lo"], start=2
            ):
                val = specs.get(key) if key != "Parameter" else param
                cell = param_map_ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = thin_border
            row_num += 1

        # Apply borders to header area too
        for row in param_map_ws.iter_rows(min_row=2, max_row=row_num-1, min_col=2, max_col=9):
            for cell in row:
                cell.border = thin_border
        self.apply_borders(param_map_ws, min_row=2, min_col=2)
        wb.save(excel_filename)
        self.show_status(f"✅ Summary sheet successfully created.")
            
    def run_histogram(self):
        try:
            wat_file = self.path_var.get()
            if not wat_file:
                self.log_error("No file selected!")
                return

            excel_filename = wat_file.replace(".wat", ".xlsx")
            wb = load_workbook(excel_filename)
            per_Wafer_ws = wb["per Wafer"]

            param_col = 7
            site_start_col = 8
            site_end_col = per_Wafer_ws.max_column - 5

            # Collect all parameters
            parameters = sorted({
                per_Wafer_ws.cell(row=row, column=param_col).value
                for row in range(2, per_Wafer_ws.max_row + 1)
                if per_Wafer_ws.cell(row=row, column=param_col).value
            })

            if not parameters:
                self.log_error("No parameters found in per Wafer sheet.")
                return

            def get_values(param):
                values, spec_hi, spec_lo, unit = [], None, None, None
                for row in range(2, per_Wafer_ws.max_row + 1):
                    if per_Wafer_ws.cell(row=row, column=param_col).value == param:
                        row_values = [
                            per_Wafer_ws.cell(row=row, column=col).value
                            for col in range(site_start_col, site_end_col + 1)
                            if isinstance(per_Wafer_ws.cell(row=row, column=col).value, (int, float))
                        ]
                        values.extend(row_values)
                        if spec_hi is None:
                            spec_hi = per_Wafer_ws.cell(row=row, column=site_end_col + 3).value
                            spec_lo = per_Wafer_ws.cell(row=row, column=site_end_col + 4).value
                            unit    = per_Wafer_ws.cell(row=row, column=site_end_col + 5).value
                return values, spec_hi, spec_lo, unit

            # Create figure with space for stats and parameter list
            fig, ax = plt.subplots(figsize=(12, 6))
            plt.subplots_adjust(right=0.65)

            # Radio buttons for parameter selection (right side)
            rax = plt.axes([0.82, 0.05, 0.15, 0.9])
            radio = RadioButtons(rax, parameters)

            # Stats panel (top right)
            stats_text_obj = fig.text(
                0.63, 0.92, "", fontsize=10,
                va="top", ha="left", family="monospace"
            )

            def plot_hist(param):
                values, spec_hi, spec_lo, unit = get_values(param)
                if not values:
                    return

                mean_val = np.mean(values)
                stdev_val = np.std(values, ddof=1)
                min_val, max_val = np.min(values), np.max(values)
                count_val = len(values)

                # Capability stats
                if spec_hi is not None and spec_lo is not None and stdev_val != 0:
                    cpk_hi = (spec_hi - mean_val) / (3 * stdev_val)
                    cpk_lo = (mean_val - spec_lo) / (3 * stdev_val)
                    cpk = min(cpk_hi, cpk_lo)
                    cp = (spec_hi - spec_lo) / (6 * stdev_val)
                else:
                    cpk_hi = cpk_lo = cpk = cp = None

                cp_str = f"{cp:.3f}" if cp is not None else "na"
                cpk_str = f"{cpk:.3f}" if cpk is not None else "na"
                cpk_hi_str = f"{cpk_hi:.3f}" if cpk_hi is not None else "na"
                cpk_lo_str = f"{cpk_lo:.3f}" if cpk_lo is not None else "na"

                # Clear and redraw histogram
                ax.clear()
                ax.hist(values, bins=20, color="#4F81BD", edgecolor="black", alpha=0.7, density=True)

                # Extend the normal curve beyond min/max by ±4σ
                x = np.linspace(mean_val - 4*stdev_val, mean_val + 4*stdev_val, 400)
                y = (1 / (stdev_val * np.sqrt(2 * np.pi))) * np.exp(-0.5 * ((x - mean_val) / stdev_val) ** 2)
                ax.plot(x, y, color="orange", linewidth=2, label="Normal Curve")

                ax.axvline(mean_val, color="green", linestyle="--", linewidth=2, label=f"Mean = {mean_val:.3f}")
                if spec_lo is not None:
                    ax.axvline(spec_lo, color="red", linestyle="--", linewidth=2, label=f"LSL = {spec_lo}")
                if spec_hi is not None:
                    ax.axvline(spec_hi, color="red", linestyle="--", linewidth=2, label=f"USL = {spec_hi}")
                ax.axvline(mean_val - 3*stdev_val, color="gray", linestyle=":", label="-3σ")
                ax.axvline(mean_val + 3*stdev_val, color="gray", linestyle=":", label="+3σ")

                ax.set_title(f"{param}")
                ax.set_xlabel(f"{unit}" if unit else "Value")
                ax.set_ylabel("Count")
                ax.legend()
                ax.grid(axis="y", linestyle="--", alpha=0.7)

                # Update stats panel text
                stats_text = f"""
    Basic Statistics
    Count: {count_val}
    Min: {min_val:.3f}
    Max: {max_val:.3f}
    Mean: {mean_val:.3f}
    Std Dev: {stdev_val:.3f}

    Specifications
    LSL: {spec_lo}
    USL: {spec_hi}
    Target: {mean_val:.3f}

    Capability Statistics
    Cp: {cp_str}
    Cpk: {cpk_str}
    Cpk Hi: {cpk_hi_str}
    Cpk Lo: {cpk_lo_str}
    """
                stats_text_obj.set_text(stats_text)
                fig.canvas.draw_idle()

                # Export the entire figure exactly as displayed
                #fig.set_size_inches(12, 6)
                #fig.savefig(f"{param}_plot_with_stats.png", dpi=300, bbox_inches="tight")

            # Connect radio button
            def on_select(label):
                plot_hist(label)

            radio.on_clicked(on_select)

            # Initial plot
            plot_hist(parameters[0])
            plt.show()

            self.log_success("Browse per Parameter Histogram....")

        except Exception as e:
            self.log_error(f"Error generating histogram: {e}")

# --- Main Entry Point ---
if __name__ == "__main__":
    root = tk.Tk()
    app = WATDataAutomation(root)
    root.mainloop()
